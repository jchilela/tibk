# sitetibl/financeiro_utils.py
"""
Utilitários para o Balanço Financeiro: cálculo de saldos por período
(semanal, mensal, trimestral, anual) do movimento geral de entradas/saídas
(banco e caixa). Inclui também a geração do relatório em Excel (.xlsx).
"""

from datetime import date
from decimal import Decimal
from io import BytesIO

from django.db.models import Sum
from django.db.models.functions import TruncWeek, TruncMonth, TruncQuarter, TruncYear

from .models import MOEDA, Entrada, Saida

PERIODO_TRUNC = {
    'semanal': TruncWeek,
    'mensal': TruncMonth,
    'trimestral': TruncQuarter,
    'anual': TruncYear,
}

PERIODOS_VALIDOS = list(PERIODO_TRUNC.keys())

MESES_PT = [
    '', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro',
]


def _label_periodo(periodo, dt):
    """Devolve um rótulo legível (pt) para o início do período `dt`."""
    if dt is None:
        return '-'
    if periodo == 'semanal':
        return f"Semana de {dt.strftime('%d/%m/%Y')}"
    if periodo == 'mensal':
        return f"{MESES_PT[dt.month]}/{dt.year}"
    if periodo == 'trimestral':
        trimestre = (dt.month - 1) // 3 + 1
        return f"{trimestre}º Trimestre/{dt.year}"
    if periodo == 'anual':
        return f"{dt.year}"
    return dt.strftime('%d/%m/%Y')


def _resolver_intervalo(params):
    """Resolve (datainicio, datafim, ano_label) a partir dos parâmetros GET."""
    ano = params.get('ano', '').strip()

    if ano and ano != 'todos':
        try:
            ano_int = int(ano)
            return date(ano_int, 1, 1), date(ano_int, 12, 31), ano
        except ValueError:
            pass

    if ano == 'todos':
        return None, None, 'todos'

    hoje = date.today()
    return date(hoje.year, 1, 1), date(hoje.year, 12, 31), str(hoje.year)


def _calcular_saldo_inicial(datainicio, moeda):
    """
    Calcula o saldo acumulado de todo o histórico ANTERIOR a `datainicio`
    (entradas gerais - saídas gerais), para que o saldo transite
    correctamente de um ano/período para o outro em vez de reiniciar em 0.
    """
    if not datainicio:
        return Decimal('0')

    filtro = {'data__lt': datainicio}
    if moeda:
        filtro['moeda'] = moeda

    total_entradas = (
        Entrada.objects.filter(**filtro).aggregate(total=Sum('valor'))['total']
        or Decimal('0')
    )
    total_saidas = (
        Saida.objects.filter(**filtro).aggregate(total=Sum('valor'))['total']
        or Decimal('0')
    )
    return total_entradas - total_saidas


def build_balanco_financeiro(params):
    """
    Constrói o contexto completo do Balanço Financeiro:
    - Entradas e Saídas gerais (banco + caixa) por período
    - Saldo do período e saldo acumulado (transitado entre períodos/anos)
    - Totais gerais
    """
    periodo = params.get('periodo', 'mensal').strip()
    if periodo not in PERIODOS_VALIDOS:
        periodo = 'mensal'
    trunc_class = PERIODO_TRUNC[periodo]

    moeda = params.get('moeda', '').strip()

    datainicio, datafim, ano = _resolver_intervalo(params)

    # ---- Filtros base -----------------------------------------------------
    filtro_movimento = {}
    if datainicio and datafim:
        filtro_movimento['data__range'] = [datainicio, datafim]
    if moeda:
        filtro_movimento['moeda'] = moeda

    # ---- Entradas gerais (banco + caixa) por período -----------------------
    entradas = (
        Entrada.objects
        .filter(**filtro_movimento)
        .annotate(periodo_dt=trunc_class('data'))
        .values('periodo_dt')
        .annotate(total=Sum('valor'))
    )

    # ---- Saídas gerais (banco + caixa) por período --------------------------
    saidas = (
        Saida.objects
        .filter(**filtro_movimento)
        .annotate(periodo_dt=trunc_class('data'))
        .values('periodo_dt')
        .annotate(total=Sum('valor'))
    )

    # ---- Consolidação por período -------------------------------------------
    periodos = {}

    def _get_bucket(periodo_dt):
        return periodos.setdefault(periodo_dt, {
            'entradas_gerais': Decimal('0'),
            'saidas_gerais': Decimal('0'),
        })

    for row in entradas:
        _get_bucket(row['periodo_dt'])['entradas_gerais'] += row['total'] or Decimal('0')
    for row in saidas:
        _get_bucket(row['periodo_dt'])['saidas_gerais'] += row['total'] or Decimal('0')

    # Saldo transitado de todo o histórico anterior ao início do período
    # filtrado — garante que o saldo acumulado nunca reinicia de um ano
    # para o outro.
    saldo_inicial = _calcular_saldo_inicial(datainicio, moeda)

    linhas = []
    saldo_acumulado = saldo_inicial
    for periodo_dt in sorted((p for p in periodos.keys() if p is not None)):
        bucket = periodos[periodo_dt]
        saldo_periodo = bucket['entradas_gerais'] - bucket['saidas_gerais']
        saldo_acumulado += saldo_periodo
        linhas.append({
            'periodo_dt': periodo_dt,
            'label': _label_periodo(periodo, periodo_dt),
            'entradas_gerais': bucket['entradas_gerais'],
            'saidas_gerais': bucket['saidas_gerais'],
            'saldo_periodo': saldo_periodo,
            'saldo_acumulado': saldo_acumulado,
        })

    # Mostrar do período mais recente para o mais antigo
    linhas_desc = list(reversed(linhas))

    total_entradas_gerais = sum((l['entradas_gerais'] for l in linhas), Decimal('0'))
    total_saidas_gerais = sum((l['saidas_gerais'] for l in linhas), Decimal('0'))
    saldo_geral = total_entradas_gerais - total_saidas_gerais
    # Saldo final = saldo transitado do histórico anterior + movimento do período filtrado.
    saldo_final = linhas[-1]['saldo_acumulado'] if linhas else saldo_inicial

    return {
        'periodo': periodo,
        'moeda': moeda,
        'moeda_choices': MOEDA,
        'ano': ano,
        'linhas': linhas_desc,
        'total_entradas_gerais': total_entradas_gerais,
        'total_saidas_gerais': total_saidas_gerais,
        'saldo_geral': saldo_geral,
        'saldo_inicial': saldo_inicial,
        'saldo_final': saldo_final,
    }


def gerar_excel_balanco_financeiro(context):
    """Gera o Balanço Financeiro em formato .xlsx (openpyxl) — layout profissional com logo e cabeçalho institucional."""
    import os
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.chart import BarChart, LineChart, Reference

    wb = Workbook()

    # ---- Estilos reutilizáveis -------------------------------------------
    COR_PRIMARIA = '1B4D3E'
    COR_PRIMARIA_CLARA = '2D6A4F'
    COR_ACENTO = 'D9EAD3'
    COR_ZEBRA = 'F0F7F4'
    COR_BORDA = 'B8C9C0'

    header_fill = PatternFill(start_color=COR_PRIMARIA, end_color=COR_PRIMARIA, fill_type='solid')
    header_font = Font(name='Calibri', color='FFFFFF', bold=True, size=11)
    title_font = Font(name='Calibri', bold=True, size=16, color=COR_PRIMARIA)
    subtitle_font = Font(name='Calibri', italic=True, size=10, color='666666')
    section_font = Font(name='Calibri', bold=True, size=12, color=COR_PRIMARIA)
    label_font = Font(name='Calibri', size=10, color='333333')
    value_font = Font(name='Calibri', size=10, color='333333')
    total_font = Font(name='Calibri', bold=True, size=11, color=COR_PRIMARIA)
    total_fill = PatternFill(start_color=COR_ACENTO, end_color=COR_ACENTO, fill_type='solid')
    zebra_fill = PatternFill(start_color=COR_ZEBRA, end_color=COR_ZEBRA, fill_type='solid')
    church_name_font = Font(name='Calibri', bold=True, size=14, color=COR_PRIMARIA)
    church_sub_font = Font(name='Calibri', size=9, color='666666', italic=True)

    thin = Side(style='thin', color=COR_BORDA)
    medium = Side(style='medium', color=COR_PRIMARIA)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_top_medium = Border(left=thin, right=thin, top=medium, bottom=thin)
    border_bottom_medium = Border(left=thin, right=thin, top=thin, bottom=medium)

    moeda_fmt = '#,##0.00 "Kz"'
    if context.get('moeda') == 'USD':
        moeda_fmt = '$#,##0.00'
    elif context.get('moeda') == 'EUR':
        moeda_fmt = '#,##0.00 "\u20ac"'

    periodo_label = {
        'semanal': 'Semanal', 'mensal': 'Mensal',
        'trimestral': 'Trimestral', 'anual': 'Anual',
    }.get(context['periodo'], context['periodo'])

    ano_label = context.get('ano', '') or 'Todos'
    moeda_label = dict(context.get('moeda_choices', [])).get(context.get('moeda', ''), 'Todas')

    # ---- Logo -----------------------------------------------------------
    logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'fotos', '2022', 'cba.png')
    logo_img = None
    if os.path.isfile(logo_path):
        try:
            logo_img = XlImage(logo_path)
            logo_img.width = 80
            logo_img.height = 80
        except Exception:
            logo_img = None

    # ====================================================================
    # Sheet 1: Resumo Executivo
    # ====================================================================
    ws = wb.active
    ws.title = 'Resumo Executivo'
    ws.sheet_view.showGridLines = False

    # Larguras de coluna
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 3
    ws.column_dimensions['E'].width = 42
    ws.column_dimensions['F'].width = 22

    # --- Cabeçalho institucional ---
    if logo_img:
        ws.add_image(logo_img, 'B2')

    ws.merge_cells('C2:F2')
    ws['C2'] = 'TERCEIRA IGREJA BAPTISTA DE LUANDA'
    ws['C2'].font = church_name_font
    ws['C2'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('C3:F3')
    ws['C3'] = 'Tabern\u00e1culo B\u00edblico da Restaura\u00e7\u00e3o \u2014 Igreja Central'
    ws['C3'].font = church_sub_font
    ws['C3'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('C4:F4')
    ws['C4'] = 'Departamento Financeiro'
    ws['C4'].font = Font(name='Calibri', size=9, color='888888')
    ws['C4'].alignment = Alignment(horizontal='left', vertical='center')

    # Linha separadora
    for col in range(2, 7):
        ws.cell(row=6, column=col).border = Border(bottom=medium)
    ws.row_dimensions[6].height = 6

    # --- Título do relatório ---
    ws.merge_cells('B8:F8')
    ws['B8'] = 'BALAN\u00c7O FINANCEIRO'
    ws['B8'].font = Font(name='Calibri', bold=True, size=18, color=COR_PRIMARIA)
    ws['B8'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[8].height = 28

    ws.merge_cells('B9:F9')
    ws['B9'] = f"Periodicidade: {periodo_label}  |  Ano: {ano_label}  |  Moeda: {moeda_label}"
    ws['B9'].font = subtitle_font
    ws['B9'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('B10:F10')
    gerado_em = datetime.now().strftime('%d/%m/%Y ' + '\u00e0s ' + '%H:%M')
    ws['B10'] = f"Gerado em: {gerado_em}"
    ws['B10'].font = Font(name='Calibri', size=9, color='999999', italic=True)
    ws['B10'].alignment = Alignment(horizontal='center', vertical='center')

    # --- KPIs principais (cards) ---
    kpi_data = [
        ('Saldo Transitado', float(context['saldo_inicial']), COR_PRIMARIA, 'FFFFFF'),
        ('Total de Entradas', float(context['total_entradas_gerais']), 'D9EAD3', COR_PRIMARIA),
        ('Total de Sa\u00eddas', float(context['total_saidas_gerais']), 'FCE4E4', '991B1B'),
        ('Saldo Final', float(context['saldo_final']), COR_PRIMARIA_CLARA, 'FFFFFF'),
    ]

    kpi_row = 13
    for idx, (label, valor, bg, fg) in enumerate(kpi_data):
        col_start = 2 + idx * 1  # B, C, E, F layout
        # Usar pares de colunas: B-C, D-E para 2x2 grid
        if idx < 2:
            r = kpi_row
            c = 2 + idx * 2  # B ou D
        else:
            r = kpi_row + 4
            c = 2 + (idx - 2) * 2  # B ou D

        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
        cell_label = ws.cell(row=r, column=c, value=label)
        cell_label.font = Font(name='Calibri', bold=True, size=9, color=fg)
        cell_label.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        cell_label.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 22

        ws.merge_cells(start_row=r + 1, start_column=c, end_row=r + 2, end_column=c + 1)
        cell_val = ws.cell(row=r + 1, column=c, value=valor)
        cell_val.font = Font(name='Calibri', bold=True, size=16, color=fg)
        cell_val.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        cell_val.alignment = Alignment(horizontal='center', vertical='center')
        cell_val.number_format = moeda_fmt
        ws.row_dimensions[r + 1].height = 20
        ws.row_dimensions[r + 2].height = 20

    # --- Tabela de indicadores detalhada ---
    row_idx = 22
    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
    ws.cell(row=row_idx, column=2, value='Indicadores Detalhados').font = section_font
    row_idx += 2

    # Cabeçalho da tabela
    ws.cell(row=row_idx, column=2, value='Indicador').font = header_font
    ws.cell(row=row_idx, column=2).fill = header_fill
    ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.cell(row=row_idx, column=2).border = border
    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=5)

    ws.cell(row=row_idx, column=6, value='Valor').font = header_font
    ws.cell(row=row_idx, column=6).fill = header_fill
    ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row_idx, column=6).border = border
    ws.row_dimensions[row_idx].height = 22
    row_idx += 1

    resumo_rows = [
        ('Saldo Transitado (Anos/Per\u00edodos Anteriores)', context['saldo_inicial'], False),
        ('Total de Entradas Gerais (Banco + Caixa)', context['total_entradas_gerais'], False),
        ('Total de Sa\u00eddas Gerais (Banco + Caixa)', context['total_saidas_gerais'], False),
        ('Movimento L\u00edquido do Per\u00edodo', context['saldo_geral'], True),
        ('Saldo Final Transitado (Acumulado)', context['saldo_final'], True),
    ]

    for i, (label, valor, is_total) in enumerate(resumo_rows):
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=5)
        cell_l = ws.cell(row=row_idx, column=2, value=label)
        cell_l.border = border
        cell_l.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        if is_total:
            cell_l.font = total_font
            cell_l.fill = total_fill
        else:
            cell_l.font = label_font
            if i % 2 == 1:
                cell_l.fill = zebra_fill

        cell_v = ws.cell(row=row_idx, column=6, value=float(valor))
        cell_v.number_format = moeda_fmt
        cell_v.border = border
        cell_v.alignment = Alignment(horizontal='right', vertical='center', indent=1)
        if is_total:
            cell_v.font = total_font
            cell_v.fill = total_fill
        else:
            cell_v.font = value_font
            if i % 2 == 1:
                cell_v.fill = zebra_fill
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1

    # --- Rodapé ---
    row_idx += 2
    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
    ws.cell(row=row_idx, column=2, value='Documento gerado automaticamente pelo sistema TIBL \u2014 Terceira Igreja Baptista de Luanda').font = Font(name='Calibri', size=8, color='AAAAAA', italic=True)
    ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')

    # ====================================================================
    # Sheet 2: Detalhe por Período
    # ====================================================================
    ws2 = wb.create_sheet('Detalhe por Per\u00edodo')
    ws2.sheet_view.showGridLines = False

    ws2.column_dimensions['A'].width = 3
    colunas = [
        ('Per\u00edodo', 24),
        ('Entradas Gerais', 20),
        ('Sa\u00eddas Gerais', 20),
        ('Saldo do Per\u00edodo', 20),
        ('Saldo Acumulado', 20),
    ]
    for col_idx, (titulo, largura) in enumerate(colunas, start=2):
        ws2.column_dimensions[get_column_letter(col_idx)].width = largura

    # Cabeçalho institucional compacto
    if logo_img:
        logo_img2 = XlImage(logo_path)
        logo_img2.width = 80
        logo_img2.height = 80
        ws2.add_image(logo_img2, 'B2')

    ws2.merge_cells('C2:F2')
    ws2['C2'] = 'TERCEIRA IGREJA BAPTISTA DE LUANDA'
    ws2['C2'].font = Font(name='Calibri', bold=True, size=12, color=COR_PRIMARIA)
    ws2['C2'].alignment = Alignment(horizontal='left', vertical='center')

    ws2.merge_cells('C3:F3')
    ws2['C3'] = 'Balan\u00e7o Financeiro \u2014 Detalhe por Per\u00edodo'
    ws2['C3'].font = subtitle_font
    ws2['C3'].alignment = Alignment(horizontal='left', vertical='center')

    ws2.merge_cells('C4:F4')
    ws2['C4'] = f"{periodo_label}  |  Ano: {ano_label}  |  Moeda: {moeda_label}"
    ws2['C4'].font = Font(name='Calibri', size=9, color='888888')
    ws2['C4'].alignment = Alignment(horizontal='left', vertical='center')

    # Linha separadora
    for col in range(2, 7):
        ws2.cell(row=6, column=col).border = Border(bottom=medium)
    ws2.row_dimensions[6].height = 6

    # Cabeçalho da tabela
    header_row = 8
    for col_idx, (titulo, _) in enumerate(colunas, start=2):
        cell = ws2.cell(row=header_row, column=col_idx, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws2.row_dimensions[header_row].height = 24

    # Dados em ordem cronológica (mais antigo primeiro)
    linhas_cronologicas = list(reversed(context['linhas']))
    for row_offset, linha in enumerate(linhas_cronologicas, start=header_row + 1):
        valores = [
            linha['label'],
            float(linha['entradas_gerais']),
            float(linha['saidas_gerais']),
            float(linha['saldo_periodo']),
            float(linha['saldo_acumulado']),
        ]
        is_zebra = (row_offset - header_row) % 2 == 0
        for col_idx, valor in enumerate(valores, start=2):
            cell = ws2.cell(row=row_offset, column=col_idx, value=valor)
            cell.border = border
            if col_idx > 2:
                cell.number_format = moeda_fmt
                cell.alignment = Alignment(horizontal='right', vertical='center', indent=1)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            cell.font = value_font
            if is_zebra:
                cell.fill = zebra_fill
        ws2.row_dimensions[row_offset].height = 20

    # Linha de totais
    total_row = header_row + 1 + len(linhas_cronologicas)
    ws2.cell(row=total_row, column=2, value='TOTAL').font = total_font
    ws2.cell(row=total_row, column=2).fill = total_fill
    ws2.cell(row=total_row, column=2).border = border
    ws2.cell(row=total_row, column=2).alignment = Alignment(horizontal='left', vertical='center', indent=1)

    for col_idx in range(3, 7):
        col_letter = get_column_letter(col_idx)
        cell = ws2.cell(row=total_row, column=col_idx)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = border
        cell.alignment = Alignment(horizontal='right', vertical='center', indent=1)
        if col_idx == 3:
            cell.value = float(context['total_entradas_gerais'])
        elif col_idx == 4:
            cell.value = float(context['total_saidas_gerais'])
        elif col_idx == 5:
            cell.value = float(context['saldo_geral'])
        elif col_idx == 6:
            cell.value = float(context['saldo_final'])
        cell.number_format = moeda_fmt
    ws2.row_dimensions[total_row].height = 24

    ws2.freeze_panes = f'B{header_row + 1}'

    # --- Gráfico de barras: Entradas vs Saídas ---
    if len(linhas_cronologicas) > 0:
        chart = BarChart()
        chart.type = 'col'
        chart.style = 2
        chart.title = 'Entradas vs Sa\u00eddas por Per\u00edodo'
        chart.y_axis.title = 'Valor'
        chart.x_axis.title = 'Per\u00edodo'
        chart.height = 10
        chart.width = 22

        data = Reference(ws2, min_col=3, max_col=4, min_row=header_row, max_row=total_row - 1)
        cats = Reference(ws2, min_col=2, min_row=header_row + 1, max_row=total_row - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws2.add_chart(chart, f'H{header_row}')

        # Gráfico de linha: Saldo Acumulado
        chart2 = LineChart()
        chart2.title = 'Evolu\u00e7\u00e3o do Saldo Acumulado'
        chart2.y_axis.title = 'Saldo'
        chart2.x_axis.title = 'Per\u00edodo'
        chart2.height = 10
        chart2.width = 22

        data2 = Reference(ws2, min_col=6, max_col=6, min_row=header_row, max_row=total_row - 1)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats)
        ws2.add_chart(chart2, f'H{header_row + 22}')

    # --- Rodapé ---
    footer_row = total_row + 3
    ws2.merge_cells(start_row=footer_row, start_column=2, end_row=footer_row, end_column=6)
    ws2.cell(row=footer_row, column=2, value='Documento gerado automaticamente pelo sistema TIBL \u2014 Terceira Igreja Baptista de Luanda').font = Font(name='Calibri', size=8, color='AAAAAA', italic=True)
    ws2.cell(row=footer_row, column=2).alignment = Alignment(horizontal='center')

    # ====================================================================
    # Configurações de impressão
    # ====================================================================
    for ws_obj in [ws, ws2]:
        ws_obj.page_setup.orientation = ws_obj.ORIENTATION_LANDSCAPE
        ws_obj.page_setup.fitToWidth = 1
        ws_obj.page_setup.fitToHeight = 0
        ws_obj.sheet_properties.pageSetUpPr.fitToPage = True
        ws_obj.page_margins.left = 0.5
        ws_obj.page_margins.right = 0.5
        ws_obj.page_margins.top = 0.7
        ws_obj.page_margins.bottom = 0.5
        ws_obj.print_options.horizontalCentered = True

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
